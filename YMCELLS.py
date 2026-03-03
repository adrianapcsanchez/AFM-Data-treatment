"""
YMPROFILE – Representative AFM Curve Selection
------------------------------------------------

This script:

1. Reads AFM force–distance .txt files
2. Allows selection of Extend or Retract segment
3. Detects contact point automatically
4. Aligns curves at contact
5. Keeps indentation region only
6. Resamples curves to equal length
7. Performs PCA dimensionality reduction
8. Clusters curves using KMeans (4 clusters)
9. Selects the most representative curve of each cluster
10. Exports raw representative curves to Excel
11. Adds representativity (%) table
12. Generates overlay plot (no offset)

Author: YMPROFILE
"""

# =========================
# Imports
# =========================

import glob
import os
import numpy as np
import pandas as pd
from scipy.interpolate import interp1d
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series


# =========================
# Function: Read AFM file
# =========================
def read_afm_file(filepath, segment_choice):
    """
    Reads AFM .txt file and extracts only the selected segment
    (extend or retract).

    Parameters:
        filepath (str): Full path to AFM file
        segment_choice (str): 'extend' or 'retract'

    Returns:
        numpy array (N x 2): position (m), force (N)
    """

    data = []
    current_segment = None

    with open(filepath, "r", errors="ignore") as fh:
        for line in fh:
            line = line.strip()

            if not line:
                continue

            # Detect segment header
            if line.startswith("#"):
                lower = line.lower()
                if "segment:" in lower:
                    if "extend" in lower:
                        current_segment = "extend"
                    elif "retract" in lower:
                        current_segment = "retract"
                continue

            # Keep only selected segment
            if current_segment != segment_choice:
                continue

            parts = line.replace(",", " ").split()
            if len(parts) >= 2:
                try:
                    data.append([float(parts[0]), float(parts[1])])
                except:
                    pass

    return np.array(data)


# =========================
# GUI Setup
# =========================

root = tk.Tk()
root.withdraw()

messagebox.showinfo("Input", "Select folder containing AFM .txt files")
input_folder = filedialog.askdirectory()

messagebox.showinfo("Output", "Select folder to save Excel file")
output_folder = filedialog.askdirectory()

# Segment selection window
segment_var = tk.StringVar(value="extend")

seg_window = tk.Toplevel()
seg_window.title("Select AFM Segment")

tk.Label(seg_window, text="Select AFM segment to analyze").pack(pady=10)
tk.Radiobutton(seg_window, text="Extend", variable=segment_var, value="extend").pack()
tk.Radiobutton(seg_window, text="Retract", variable=segment_var, value="retract").pack()

tk.Button(seg_window, text="OK", command=seg_window.destroy).pack(pady=10)
root.wait_window(seg_window)

segment_choice = segment_var.get()

# Ask Excel file name
excel_name = simpledialog.askstring(
    "File Name",
    "Enter Excel name (without prefix):"
)

if not excel_name:
    messagebox.showerror("Error", "No file name provided.")
    exit()

excel_filename = f"YMPROFILE_{excel_name}.xlsx"


# =========================
# Collect files (remove duplicates)
# =========================

files = glob.glob(os.path.join(input_folder, "**/*.txt"), recursive=True)
files = list({os.path.abspath(f) for f in files})

if len(files) == 0:
    messagebox.showerror("Error", "No .txt files found.")
    exit()


# =========================
# Read and preprocess curves
# =========================

processed_curves = []
filepaths_valid = []

for filepath in files:

    arr = read_afm_file(filepath, segment_choice)

    if len(arr) < 100:
        continue

    position_nm = arr[:, 0] * 1e9
    force_pN = arr[:, 1] * 1e12

    # Ensure consistent orientation
    if position_nm[0] > position_nm[-1]:
        position_nm = position_nm[::-1]
        force_pN = force_pN[::-1]

    # Smooth force (rolling mean)
    force_smooth = pd.Series(force_pN).rolling(5, center=True).mean()
    force_smooth = force_smooth.bfill().ffill().values

    # Contact point detection (5% of max force)
    threshold = 0.05 * np.max(force_smooth)
    contact_idx = np.argmax(force_smooth > threshold)

    contact_position = position_nm[contact_idx]

    # Compute indentation
    indentation = position_nm - contact_position

    # Keep only indentation region
    mask = (indentation >= 0) & (force_smooth >= 0)

    indentation = indentation[mask]
    force_indent = force_smooth[mask]

    if len(force_indent) > 50:
        processed_curves.append((indentation, force_indent))
        filepaths_valid.append(filepath)

if len(processed_curves) < 4:
    messagebox.showerror("Error", "Not enough valid indentation curves.")
    exit()


# =========================
# Resample curves to equal length
# =========================

target_length = 400
resampled = []

for indentation, force in processed_curves:
    x_original = np.linspace(0, 1, len(force))
    x_target = np.linspace(0, 1, target_length)
    interpolator = interp1d(x_original, force, kind="linear")
    resampled.append(interpolator(x_target))

X = np.array(resampled)


# =========================
# PCA + KMeans clustering
# =========================

pca = PCA(n_components=5)
Xp = pca.fit_transform(X)

kmeans = KMeans(n_clusters=4, random_state=0)
kmeans.fit(Xp)


# =========================
# Select representative curve of each cluster
# =========================

selected_indices = []
cluster_sizes = []
cluster_percentages = []

total_curves = len(processed_curves)

for i in range(4):

    cluster_points = np.where(kmeans.labels_ == i)[0]
    count = len(cluster_points)

    cluster_sizes.append(count)
    cluster_percentages.append((count / total_curves) * 100)

    centroid = kmeans.cluster_centers_[i]
    distances = np.linalg.norm(Xp[cluster_points] - centroid, axis=1)

    representative_index = cluster_points[np.argmin(distances)]
    selected_indices.append(representative_index)


# =========================
# Export RAW representative curves
# =========================

columns = {}
max_length = 0

for idx in selected_indices:

    filepath = filepaths_valid[idx]
    filename = os.path.basename(filepath)

    arr = read_afm_file(filepath, segment_choice)

    position_nm = arr[:, 0] * 1e9
    force_pN = arr[:, 1] * 1e12

    max_length = max(max_length, len(position_nm))

    columns[f"{filename}_Position_nm"] = list(position_nm)
    columns[f"{filename}_Force_pN"] = list(force_pN)

# Pad columns
for key in columns:
    if len(columns[key]) < max_length:
        columns[key].extend([np.nan] * (max_length - len(columns[key])))

df = pd.DataFrame(columns)

excel_path = os.path.join(output_folder, excel_filename)
df.to_excel(excel_path, index=False)


# =========================
# Add Excel Chart (overlay, no offset)
# =========================

wb = load_workbook(excel_path)
ws = wb.active

chart = ScatterChart()
chart.title = f"Representative AFM Curves ({segment_choice})"
chart.x_axis.title = "Position (nm)"
chart.y_axis.title = "Force (pN)"
chart.height = 15
chart.width = 25

for i in range(4):

    x_col = 1 + i * 2
    y_col = 2 + i * 2

    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=max_length+1)
    yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=max_length+1)

    series = Series(yvalues, xvalues, title=ws.cell(row=1, column=y_col).value)
    chart.series.append(series)

ws.add_chart(chart, "K2")


# =========================
# Add representativity table
# =========================

start_col = 10

ws.cell(row=1, column=start_col).value = "Representative Curve"
ws.cell(row=1, column=start_col+1).value = "Similar Curves (n)"
ws.cell(row=1, column=start_col+2).value = "Percentage (%)"

for i in range(4):
    filename = os.path.basename(filepaths_valid[selected_indices[i]])
    ws.cell(row=i+2, column=start_col).value = filename
    ws.cell(row=i+2, column=start_col+1).value = cluster_sizes[i]
    ws.cell(row=i+2, column=start_col+2).value = round(cluster_percentages[i], 2)

wb.save(excel_path)


print("Total curves used in clustering:", total_curves)
print("Sum of cluster sizes:", sum(cluster_sizes))

messagebox.showinfo("Completed", f"File saved at:\n{excel_path}")