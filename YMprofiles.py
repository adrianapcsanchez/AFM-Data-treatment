import glob
import os
import tkinter as tk
from tkinter import filedialog, messagebox

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from scipy.interpolate import interp1d
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA


# =========================
# Folder selection
# =========================
root = tk.Tk()
root.withdraw()

messagebox.showinfo("Input", "Select the folder containing AFM curves")
input_folder = filedialog.askdirectory()

messagebox.showinfo("Output", "Select where the Excel file will be saved")
output_folder = filedialog.askdirectory()


# =========================
# Search files
# =========================
files = glob.glob(os.path.join(input_folder, "**/*.txt"), recursive=True)
files += glob.glob(os.path.join(input_folder, "**/*.TXT"), recursive=True)

print("Files found:", len(files))

curves = []
names = []


# =========================
# Robust reading
# =========================
for f in files:
    data = []
    with open(f, "r", errors="ignore") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue

            parts = line.replace(",", " ").split()
            if len(parts) >= 2:
                try:
                    data.append([float(parts[0]), float(parts[1])])
                except ValueError:
                    pass

    arr = np.array(data)

    if len(arr) > 100:
        curves.append(arr)
        names.append(os.path.basename(f))

print("Valid curves:", len(curves))

if len(curves) < 4:
    messagebox.showerror("Error", "Less than 4 valid curves found.")
    raise SystemExit


# =========================
# Representative selection
# =========================
target_len = 500
resampled = []

for c in curves:
    x = np.linspace(0, 1, len(c))
    xi = np.linspace(0, 1, target_len)
    f_interp = interp1d(x, c[:, 1], kind="linear")
    resampled.append(f_interp(xi))

X = np.array(resampled)

pca = PCA(n_components=5)
Xp = pca.fit_transform(X)

kmeans = KMeans(n_clusters=4, random_state=0).fit(Xp)

selected_idx = []
for i in range(4):
    cluster_points = np.where(kmeans.labels_ == i)[0]
    centroid = kmeans.cluster_centers_[i]
    distances = np.linalg.norm(Xp[cluster_points] - centroid, axis=1)
    selected_idx.append(cluster_points[np.argmin(distances)])

selected_files = [names[i] for i in selected_idx]

print("Selected curves:")
for f in selected_files:
    print(f)


# =========================
# Create single table
# =========================
all_columns = {}
max_len = 0

offset_multiplier = 1.2
offset_value = 0

for f in selected_files:
    filepath = None
    for full in files:
        if full.endswith(f):
            filepath = full
            break

    pos_list = []
    force_list = []

    with open(filepath, "r", errors="ignore") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue

            parts = line.replace(",", " ").split()
            if len(parts) >= 2:
                try:
                    pos_nm = float(parts[0]) * 1e9
                    force_pN = float(parts[1]) * 1e12
                    pos_list.append(pos_nm)
                    force_list.append(force_pN + offset_value)
                except ValueError:
                    pass

    max_len = max(max_len, len(pos_list))

    all_columns[f"{f}_Position_nm"] = pos_list
    all_columns[f"{f}_Force_pN"] = force_list

    offset_value += max(force_list) * offset_multiplier


# Adjust column sizes
for key in all_columns:
    if len(all_columns[key]) < max_len:
        all_columns[key].extend([np.nan] * (max_len - len(all_columns[key])))

df = pd.DataFrame(all_columns)


# =========================
# Save Excel
# =========================
excel_path = os.path.join(output_folder, "YMProfiles.xlsx")
df.to_excel(excel_path, index=False)


# =========================
# Create Excel chart
# =========================
wb = load_workbook(excel_path)
ws = wb.active

chart = ScatterChart()
chart.title = "Representative AFM Curves"
chart.x_axis.title = "Position (nm)"
chart.y_axis.title = "Force (pN) - Offset"
chart.height = 15
chart.width = 25

for i in range(4):
    x_col = 1 + i * 2
    y_col = 2 + i * 2

    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=max_len + 1)
    yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=max_len + 1)

    series = Series(yvalues, xvalues, title=ws.cell(row=1, column=y_col).value)
    chart.series.append(series)

ws.add_chart(chart, "K2")

wb.save(excel_path)

messagebox.showinfo("Completed", f"File saved with chart at:\n{excel_path}")
